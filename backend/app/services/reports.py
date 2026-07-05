import io
import datetime
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.models import Machine, Alert, MaintenanceTask, InventoryItem

def generate_pdf_report(db: Session) -> io.BytesIO:
    """
    Generates a professional PDF report summarizing the current state of the plant.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom colors and styles
    navy = colors.HexColor("#1e293b")
    blue_accent = colors.HexColor("#0284c7")
    gray_light = colors.HexColor("#f8fafc")
    text_dark = colors.HexColor("#0f172a")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=navy,
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=25
    )
    
    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=blue_accent,
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyDark',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=text_dark,
        leading=12
    )

    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # 1. Header Section
    story.append(Paragraph("SMART MANUFACTURING PLATFORM", title_style))
    story.append(Paragraph(f"Predictive Maintenance Executive Report | Generated: {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # 2. Key Performance Indicators (KPIs)
    machines = db.query(Machine).all()
    alerts = db.query(Alert).filter(Alert.acknowledged == False).all()
    tasks = db.query(MaintenanceTask).filter(MaintenanceTask.status.in_(["scheduled", "in_progress"])).all()
    
    avg_health = sum([m.health_score for m in machines]) / len(machines) if machines else 0
    avg_oee = sum([m.oee for m in machines]) / len(machines) if machines else 0
    
    kpi_data = [
        [
            Paragraph(f"<b>Total Machines:</b> {len(machines)}", body_style),
            Paragraph(f"<b>Active Alerts:</b> {len(alerts)}", body_style),
            Paragraph(f"<b>Avg Health:</b> {round(avg_health, 1)}%", body_style)
        ],
        [
            Paragraph(f"<b>Scheduled Tasks:</b> {len(tasks)}", body_style),
            Paragraph(f"<b>Avg OEE:</b> {round(avg_oee, 1)}%", body_style),
            Paragraph("", body_style) # Spacer
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[180, 180, 180])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), gray_light),
        ('PADDING', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#e2e8f0")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#f1f5f9")),
    ]))
    
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # 3. Machine Fleet Summary Table
    story.append(Paragraph("Machine Fleet Performance", h2_style))
    
    machine_headers = [
        Paragraph("ID", header_style),
        Paragraph("Machine Name", header_style),
        Paragraph("Type", header_style),
        Paragraph("Health", header_style),
        Paragraph("OEE", header_style),
        Paragraph("Predicted RUL", header_style),
        Paragraph("Status", header_style)
    ]
    
    m_table_content = [machine_headers]
    for m in machines:
        status_color = "#22c55e"  # healthy green
        if m.status == "warning":
            status_color = "#f59e0b"  # warning amber
        elif m.status == "critical":
            status_color = "#ef4444"  # critical red
            
        m_table_content.append([
            Paragraph(str(m.id), body_style),
            Paragraph(m.name, body_style),
            Paragraph(m.type.upper(), body_style),
            Paragraph(f"{m.health_score}%", body_style),
            Paragraph(f"{m.oee}%", body_style),
            Paragraph(f"{m.rul_hours} hrs", body_style),
            Paragraph(f"<font color='{status_color}'><b>{m.status.upper()}</b></font>", body_style)
        ])
        
    m_table = Table(m_table_content, colWidths=[25, 160, 75, 55, 55, 80, 80])
    m_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, gray_light]),
    ]))
    story.append(m_table)
    story.append(Spacer(1, 20))
    
    # 4. Alerts and Schedule details (Page Break for page alignment if list is long)
    story.append(Paragraph("Unresolved Industrial Alerts", h2_style))
    if not alerts:
        story.append(Paragraph("No active unresolved alerts recorded in the last 24 hours.", body_style))
    else:
        alert_headers = [
            Paragraph("Time", header_style),
            Paragraph("Machine", header_style),
            Paragraph("Severity", header_style),
            Paragraph("Alert Message", header_style)
        ]
        a_table_content = [alert_headers]
        for a in alerts[:15]:
            sev_color = "#ef4444" if a.severity == "critical" else "#f59e0b"
            a_table_content.append([
                Paragraph(a.timestamp.strftime('%H:%M:%S'), body_style),
                Paragraph(a.machine.name, body_style),
                Paragraph(f"<font color='{sev_color}'><b>{a.severity.upper()}</b></font>", body_style),
                Paragraph(a.message, body_style)
            ])
        a_table = Table(a_table_content, colWidths=[70, 120, 65, 275])
        a_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, gray_light]),
        ]))
        story.append(a_table)
        
    story.append(Spacer(1, 20))
    
    # 5. Maintenance Tickets
    story.append(Paragraph("Active Scheduled Maintenance Tickets", h2_style))
    if not tasks:
        story.append(Paragraph("No maintenance operations scheduled currently.", body_style))
    else:
        task_headers = [
            Paragraph("Machine", header_style),
            Paragraph("Maintenance Task", header_style),
            Paragraph("Priority", header_style),
            Paragraph("Technician", header_style),
            Paragraph("Date", header_style),
            Paragraph("Downtime", header_style)
        ]
        t_table_content = [task_headers]
        for t in tasks[:15]:
            t_table_content.append([
                Paragraph(t.machine.name, body_style),
                Paragraph(t.title, body_style),
                Paragraph(t.priority.upper(), body_style),
                Paragraph(t.assigned_engineer or "Unassigned", body_style),
                Paragraph(t.scheduled_date.strftime('%Y-%m-%d'), body_style),
                Paragraph(f"{t.estimated_downtime_hours} hrs", body_style)
            ])
        t_table = Table(t_table_content, colWidths=[100, 160, 60, 85, 75, 50])
        t_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, gray_light]),
        ]))
        story.append(t_table)
        
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(db: Session) -> io.BytesIO:
    """
    Generates a multi-sheet, beautifully formatted Excel spreadsheet.
    """
    wb = Workbook()
    
    # Styles config
    font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_navy = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    fill_header = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    fill_light = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Sheet 1: Fleet Status
    ws1 = wb.active
    ws1.title = "Fleet Performance"
    
    # Title Block
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "MACHINE FLEET STATUS REPORT"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 40
    
    # Headers
    headers1 = ["Machine ID", "Name", "Type", "Health Score (%)", "OEE (%)", "Predicted RUL (Hours)", "Status"]
    ws1.append([]) # Blank row
    ws1.append(headers1)
    ws1.row_dimensions[3].height = 25
    
    for col_idx in range(1, 8):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    machines = db.query(Machine).all()
    for m in machines:
        ws1.append([m.id, m.name, m.type.upper(), m.health_score, m.oee, m.rul_hours, m.status.upper()])
        
    # Formatting body rows
    for r in range(4, ws1.max_row + 1):
        ws1.row_dimensions[r].height = 20
        # Color Status cell based on warning state
        status_val = ws1.cell(row=r, column=7).value
        if status_val == "CRITICAL":
            ws1.cell(row=r, column=7).font = Font(name="Calibri", size=11, bold=True, color="990000")
        elif status_val == "WARNING":
            ws1.cell(row=r, column=7).font = Font(name="Calibri", size=11, bold=True, color="B45309")
        else:
            ws1.cell(row=r, column=7).font = Font(name="Calibri", size=11, bold=True, color="047857")
            
        for c in range(1, 8):
            cell = ws1.cell(row=r, column=c)
            cell.border = border_thin
            if r % 2 == 1:
                cell.fill = fill_light
                
    # Sheet 2: Unresolved Alerts
    ws2 = wb.create_sheet(title="Active Alerts")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "ACTIVE INDUSTRIAL TELEMETRY ALERTS"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_navy
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 40
    
    headers2 = ["Alert ID", "Machine ID", "Machine Name", "Severity", "Message"]
    ws2.append([])
    ws2.append(headers2)
    ws2.row_dimensions[3].height = 25
    
    for col_idx in range(1, 6):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    alerts = db.query(Alert).filter(Alert.acknowledged == False).all()
    for a in alerts:
        ws2.append([a.id, a.machine_id, a.machine.name, a.severity.upper(), a.message])
        
    for r in range(4, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 20
        # Color code severity
        sev_val = ws2.cell(row=r, column=4).value
        if sev_val == "CRITICAL":
            ws2.cell(row=r, column=4).font = Font(name="Calibri", size=11, bold=True, color="990000")
        else:
            ws2.cell(row=r, column=4).font = Font(name="Calibri", size=11, bold=True, color="B45309")
            
        for c in range(1, 6):
            cell = ws2.cell(row=r, column=c)
            cell.border = border_thin
            if r % 2 == 1:
                cell.fill = fill_light

    # Sheet 3: Maintenance Schedule
    ws3 = wb.create_sheet(title="Maintenance Schedule")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "SCHEDULED PREVENTATIVE MAINTENANCE"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_navy
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 40
    
    headers3 = ["Task ID", "Machine Name", "Task Title", "Priority", "Technician", "Scheduled Date", "Downtime (Hours)"]
    ws3.append([])
    ws3.append(headers3)
    ws3.row_dimensions[3].height = 25
    
    for col_idx in range(1, 8):
        cell = ws3.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    tasks = db.query(MaintenanceTask).all()
    for t in tasks:
        ws3.append([
            t.id, 
            t.machine.name, 
            t.title, 
            t.priority.upper(), 
            t.assigned_engineer or "UNASSIGNED", 
            t.scheduled_date.strftime('%Y-%m-%d %H:%M'), 
            t.estimated_downtime_hours
        ])
        
    for r in range(4, ws3.max_row + 1):
        ws3.row_dimensions[r].height = 20
        for c in range(1, 8):
            cell = ws3.cell(row=r, column=c)
            cell.border = border_thin
            if r % 2 == 1:
                cell.fill = fill_light

    # Sheet 4: Inventory
    ws4 = wb.create_sheet(title="Inventory Optimization")
    ws4.merge_cells("A1:G1")
    ws4["A1"] = "SPARE PARTS INVENTORY & REORDER SUGGESTIONS"
    ws4["A1"].font = font_title
    ws4["A1"].fill = fill_navy
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 40
    
    headers4 = ["Part ID", "Part Name", "Part Number", "Stock Level", "Min Stock Level", "Unit Cost ($)", "Status"]
    ws4.append([])
    ws4.append(headers4)
    ws4.row_dimensions[3].height = 25
    
    for col_idx in range(1, 8):
        cell = ws4.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    inventory = db.query(InventoryItem).all()
    for item in inventory:
        reorder_status = "RESTOCK" if item.stock_level < item.minimum_stock else "ADEQUATE"
        ws4.append([item.id, item.name, item.part_number, item.stock_level, item.minimum_stock, item.unit_cost, reorder_status])
        
    for r in range(4, ws4.max_row + 1):
        ws4.row_dimensions[r].height = 20
        status_val = ws4.cell(row=r, column=7).value
        if status_val == "RESTOCK":
            ws4.cell(row=r, column=7).font = Font(name="Calibri", size=11, bold=True, color="990000")
            ws4.cell(row=r, column=7).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        else:
            ws4.cell(row=r, column=7).font = Font(name="Calibri", size=11, bold=True, color="047857")
            
        for c in range(1, 8):
            cell = ws4.cell(row=r, column=c)
            cell.border = border_thin
            if r % 2 == 1 and status_val != "RESTOCK":
                cell.fill = fill_light
                
    # Auto-adjust column widths across all worksheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Find the longest value in the column
            for cell in col:
                if cell.row == 1: # Skip title row calculation
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
